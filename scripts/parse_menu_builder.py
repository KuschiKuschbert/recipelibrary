#!/usr/bin/env python3
"""
Parse Riviera Menu Builder.xlsx → reports/reference_sheet_extract.json
                                + reports/workbook_catalogue.md

Reads all sheets, dumps cell values, and classifies each tab:
  - live_stock_card   : single-ingredient cost card
  - live_dish_card    : real dish with ingredients/costs
  - stock_list        : aggregated stock / par / storage list
  - dashboard         : formula-heavy summary / reporting tab
  - template_demo     : old template / placeholder / demo data
  - engineering       : tool / audit / migration ledger

Usage:
    python3 scripts/parse_menu_builder.py [--xlsx PATH]
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl not installed. Run: pip3 install openpyxl")

ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "reports"
REPORTS.mkdir(exist_ok=True)

DEMO_SIGNALS = {
    "pork thingies", "beef thingamajigs", "spanish flu", "chinese wuhan thing",
    "kburger", "ksteak", "flap flaps", "chicken tiddies", "pork thingies",
}
DASH_TABS = {
    "Dashboard", "Menu Performance", "Volume GP", "Engineering",
    "Recipe Migration Ledger", "Stock Match Audit P1", "Guide",
    "Dishes Dashboard",
}
STOCK_LIST_TABS = {"Stock List", "Stock by Storage", "Template"}
CATERING_DISH_TABS = {
    "Breads", "Starters", "Oysters", "Salads", "Pizzas",
    "Lunch", "Chef Selection", "Italian Long Lunch",
    "Steaks + Grill", "Mains", "Sides", "Toppers",
    "Kids Meals", "Desserts",
}
# Per-ingredient stock tabs match this pattern (all-caps + underscores usually)
STOCK_CARD_RE = re.compile(r'^[A-Z][A-Z0-9_\- ]{3,}$')


def sheet_to_rows(ws, max_rows: int = 200, max_cols: int = 30) -> list[list]:
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True), 1):
        if i > max_rows:
            break
        rows.append([v for v in row])
    return rows


def classify_tab(name: str, rows: list[list]) -> str:
    if name in DASH_TABS:
        return "dashboard"
    if name in STOCK_LIST_TABS:
        return "stock_list"
    if name in CATERING_DISH_TABS:
        return "live_dish_card"

    # Check if it looks like a per-ingredient stock card
    if STOCK_CARD_RE.match(name):
        return "live_stock_card"

    # Look for demo signals in cell values
    flat_vals = {str(v).strip().lower() for row in rows[:30] for v in row if v is not None}
    if flat_vals & DEMO_SIGNALS:
        return "template_demo"

    return "unknown"


def sample_rows(rows: list[list], n: int = 5) -> list[list]:
    """Return first n non-empty rows."""
    out = []
    for row in rows:
        if any(v is not None for v in row):
            out.append([str(v) if v is not None else "" for v in row])
        if len(out) >= n:
            break
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--xlsx",
        default=str(Path.home() / "Downloads" / "Riviera Menu Builder.xlsx"),
    )
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")

    print(f"Loading {xlsx_path} …", flush=True)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    print(f"  {len(sheet_names)} sheets found", flush=True)

    extract: dict = {}
    catalogue: list[dict] = []

    for name in sheet_names:
        ws = wb[name]
        rows = sheet_to_rows(ws, max_rows=300, max_cols=40)
        tab_class = classify_tab(name, rows)

        # Count non-empty rows and cols
        non_empty = sum(1 for row in rows if any(v is not None for v in row))

        # Find header row (first row with >= 2 non-None values)
        header: list[str] = []
        for row in rows[:10]:
            vals = [str(v) if v is not None else "" for v in row]
            if sum(1 for v in vals if v.strip()) >= 2:
                header = [v.strip() for v in vals]
                break

        samples = sample_rows(rows, 8)

        entry = {
            "class": tab_class,
            "non_empty_rows": non_empty,
            "header": header,
            "sample_rows": samples,
            "all_rows": rows,
        }
        extract[name] = entry

        catalogue.append({
            "tab": name,
            "class": tab_class,
            "non_empty_rows": non_empty,
            "header": header,
            "sample_rows": samples,
        })
        print(f"  [{tab_class:18s}] {name} ({non_empty} non-empty rows)")

    wb.close()

    # Write extract (without all_rows to keep it manageable; write a separate full file)
    extract_slim = {
        k: {kk: vv for kk, vv in v.items() if kk != "all_rows"}
        for k, v in extract.items()
    }
    out_json = REPORTS / "reference_sheet_extract.json"
    out_json.write_text(json.dumps(extract_slim, indent=2, default=str), encoding="utf-8")
    print(f"\nWrote {out_json}")

    # Write full extract (all row data) for later phases
    out_full = REPORTS / "reference_sheet_extract_full.json"
    out_full.write_text(json.dumps(extract, indent=2, default=str), encoding="utf-8")
    print(f"Wrote {out_full}")

    # Write markdown catalogue
    lines = [
        "# Riviera Menu Builder — Workbook Catalogue",
        "",
        f"Source: `{xlsx_path.name}`  ",
        f"Sheets: {len(sheet_names)}",
        "",
        "## Classification legend",
        "| Class | Meaning |",
        "|---|---|",
        "| `live_stock_card` | Single-ingredient cost/recipe card |",
        "| `live_dish_card` | Real dish with ingredients/costs |",
        "| `stock_list` | Aggregated stock, par or storage list |",
        "| `dashboard` | Formula-heavy summary, engineering or reporting |",
        "| `template_demo` | Old template / placeholder / demo data |",
        "| `unknown` | Needs manual review |",
        "",
        "## Tab summary",
        "",
    ]

    by_class: dict[str, list] = {}
    for row in catalogue:
        by_class.setdefault(row["class"], []).append(row)

    for cls, tabs in sorted(by_class.items()):
        lines.append(f"### {cls} ({len(tabs)} tabs)")
        for t in tabs:
            header_str = " · ".join(h for h in t["header"][:6] if h)
            lines.append(f"- **{t['tab']}** — {t['non_empty_rows']} rows")
            if header_str:
                lines.append(f"  - columns: `{header_str}`")
            if t["sample_rows"]:
                first = t["sample_rows"][0]
                first_str = " | ".join(str(v) for v in first[:6] if str(v).strip())
                if first_str:
                    lines.append(f"  - first data row: {first_str}")
        lines.append("")

    lines.append("## Action plan")
    lines.append("")
    lines.append("**DO NOT change any app data until this catalogue has been reviewed.**")
    lines.append("")
    lines.append("- `live_stock_card` tabs → Phase 1: reconcile ingredient names/units into builtins.json")
    lines.append("- `stock_list` tabs → Phase 2: update par levels + order-list storage zones")
    lines.append("- `live_dish_card` tabs → Phase 3: import only dishes that map to catering packages")
    lines.append("- `dashboard` tabs → read-only reference; no import")
    lines.append("- `template_demo` tabs → skip entirely")
    lines.append("- `unknown` tabs → requires manual decision")

    out_md = REPORTS / "workbook_catalogue.md"
    out_md.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {out_md}")
    print("\nPhase 0 complete. Review the catalogue before continuing.")


if __name__ == "__main__":
    main()
