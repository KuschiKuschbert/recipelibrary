#!/usr/bin/env python3
"""Build Riviera's GitHub recipe-data bundle.

GitHub is canonical for structured recipe data. The verified 2026-07-08
ChatGPT Riviera Project snapshot is retained as historical provenance, while
Google Drive remains the editable operational master and ChatGPT consumes a
read-optimised published release. The fixed July 8 tapas/canape overlay and
later approved house standards are represented in the structured catalog.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import zipfile
from collections import defaultdict
from html import unescape
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "riviera_sources" / "chatgpt_project_sources_2026-07-08"
OUTPUT_DIR = ROOT / "riviera_sources" / "current"
RECIPE_CATALOG_PATH = OUTPUT_DIR / "Riviera_Recipe_Catalog_Source_Of_Truth_2026-07-08.json"
PACKAGES_PATH = ROOT / "riviera_data" / "function_packages.json"

SOURCE_OF_TRUTH_PATH = OUTPUT_DIR / "Riviera_Source_Of_Truth_2026-07-08.md"
TAPAS_OVERLAY_PATH = OUTPUT_DIR / "Riviera_Tapas_House_Standards_Overlay_2026-07-08.md"
CURRENT_STANDARDS_ADDITIONS_PATH = OUTPUT_DIR / "Riviera_Current_House_Standards_Additions_2026-07-27.md"
MANIFEST_PATH = OUTPUT_DIR / "manifest.json"
ACTIVE_RELEASE_ID = "RIV-KNOWLEDGE-2026-07-27-V13"

JULY_8_OVERLAY_IDS = [
    "arancini",
    "calamari",
    "oysters-kilpatrick",
    "veal-meatballs",
    "chicken-skewer",
    "chorizo-potatoes",
    "lamb-cutlet",
    "fish-slider",
    "romesco",
    "lemon-dill-aioli",
    "lemon-thyme-aioli",
    "vodka-sauce",
    "riviera-emulsion",
    "whipped-butter",
    "camembert-cigars",
    "beef-kofta",
]
CURRENT_HOUSE_STANDARD_IDS = [
    *JULY_8_OVERLAY_IDS,
    "peach-tartare",
    "house-scones",
    "potato-pave",
    "baklava-cheesecake",
    "house-focaccia",
    "burnt-butter-mash",
]

SOURCE_FILES: list[dict[str, str]] = [
    {"file": "Baclava-Cheesecake.txt", "kind": "text", "liveName": "Baclava-Cheesecake.txt"},
    {"file": "Riviera_Sunday_Tapas_Pull_Matrix_SOP_Addendum_v1_2026-06-16.md", "kind": "text", "liveName": "Riviera_Sunday_Tapas_Pull_Matrix_SOP_Addendum_v1_2026-06-16.md"},
    {
        "file": "Riviera_Sunday_Tapas_Pull_Matrix_A4_Kitchen_Sheet_v1_2026-06-16.txt",
        "kind": "text",
        "liveName": "Riviera-Sunday-Tapas-Pull-Matrix-—-A4-Kitchen-Sheet-v1.txt",
    },
    {"file": "kitchen_pull_matrix.pdf", "kind": "pdf", "extract": "extracted/kitchen_pull_matrix.pdf.extracted.md", "liveName": "kitchen_pull_matrix.pdf"},
    {"file": "Riviera_SOP_Master_Index_v10_2026-06-16.md", "kind": "text", "liveName": "Riviera_SOP_Master_Index_v10_2026-06-16.md"},
    {"file": "Riviera_Kitchen_Production_Harness_Index_v5_2026-06-16.md", "kind": "text", "liveName": "Riviera_Kitchen_Production_Harness_Index_v5_2026-06-16.md"},
    {"file": "Riviera_Count_Ordering_Harness_v3_2026-06-16.md", "kind": "text", "liveName": "Riviera_Count_Ordering_Harness_v3_2026-06-16.md"},
    {"file": "Riviera_Component_Module_Library_v3_2026-06-16.md", "kind": "text", "liveName": "Riviera_Component_Module_Library_v3_2026-06-16.md"},
    {"file": "Riviera_Real_Event_Test_Pack_v2_2026-06-16.md", "kind": "text", "liveName": "Riviera_Real_Event_Test_Pack_v2_2026-06-16.md"},
    {"file": "Riviera_Production_Sheet_Template_Library_v2_2026-06-16.md", "kind": "text", "liveName": "Riviera_Production_Sheet_Template_Library_v2_2026-06-16.md"},
    {"file": "MYO bars and Buffets.pdf", "kind": "pdf", "extract": "extracted/MYO bars and Buffets.pdf.extracted.md", "liveName": "MYO bars and Buffets.pdf"},
    {"file": "Updated_Riviera_instructions_2026-06-09_extracted.md", "kind": "text", "liveName": "Updated Riviera instructions"},
    {"file": "foodpairing_condensed_riviera_reference.md", "kind": "text", "liveName": "foodpairing_condensed_riviera_reference.md"},
    {"file": "Riviera_Weekly_Order_Workflow_v1_2026-06-09.md", "kind": "text", "liveName": "Riviera_Weekly_Order_Workflow_v1_2026-06-09.md"},
    {"file": "Riviera_Package_Source_Digest_v1_2026-06-09.md", "kind": "text", "liveName": "Riviera_Package_Source_Digest_v1_2026-06-09.md"},
    {"file": "Riviera_Order_Template_v1_2026-06-09.md", "kind": "text", "liveName": "Riviera_Order_Template_v1_2026-06-09.md"},
    {"file": "Riviera_Canonical_Recipe_Bank_v1_2026-06-08.md", "kind": "text", "liveName": "Riviera_Canonical_Recipe_Bank_v1_2026-06-08.md"},
    {"file": "Riviera_Supplier_Ordering_Translator_v1_2026-06-08.md", "kind": "text", "liveName": "Riviera_Supplier_Ordering_Translator_v1_2026-06-08.md"},
    {"file": "Riviera_Seasoning_Palette_v2_2026-06-08.md", "kind": "text", "liveName": "Riviera_Seasoning_Palette_v2_2026-06-08.md"},
    {"file": "Recipes for Prep Chef.docx", "kind": "docx", "extract": "extracted/Recipes for Prep Chef.docx.extracted.md", "liveName": "Recipes for Prep Chef.docx"},
    {
        "file": "Olive+Green+on+White+Background.webp.source-record.md",
        "kind": "image-source-record",
        "liveName": "Olive+Green+on+White+Background.webp",
    },
    {"file": "Bidfood_Item_List.pdf", "kind": "pdf", "extract": "extracted/Bidfood_Item_List.pdf.extracted.md", "liveName": "Bidfood_Item_List.pdf"},
    {
        "file": "OrderForm11579311 - 1098368202605251257349179710.xlsx",
        "kind": "xlsx",
        "extract": "extracted/OrderForm11579311 - 1098368202605251257349179710.xlsx.extracted.md",
        "liveName": "OrderForm11579311 - 1098368202605251257349179710.xlsx",
    },
]


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def load_recipe_catalog() -> list[dict[str, Any]]:
    if not RECIPE_CATALOG_PATH.exists():
        raise SystemExit(f"Missing structured Riviera recipe catalog: {RECIPE_CATALOG_PATH.relative_to(ROOT)}")
    payload = load_json(RECIPE_CATALOG_PATH)
    if not isinstance(payload, dict) or not isinstance(payload.get("recipes"), list):
        raise SystemExit(f"{RECIPE_CATALOG_PATH.relative_to(ROOT)} must be an object with a recipes list")
    if payload.get("releaseId") != ACTIVE_RELEASE_ID:
        raise SystemExit(
            f"{RECIPE_CATALOG_PATH.relative_to(ROOT)} releaseId must be {ACTIVE_RELEASE_ID}"
        )
    return payload["recipes"]


def clean(value: Any) -> str:
    return " ".join(str(value or "").split())


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    h.update(path.read_bytes())
    return h.hexdigest()


def line_count(text: str) -> int:
    return text.count("\n") + (0 if text.endswith("\n") else 1)


def write_if_changed(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists() and path.read_text(encoding="utf-8") == content:
        return
    path.write_text(content, encoding="utf-8")


def extract_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    lines = [f"# Extracted Text: {path.name}", "", f"Pages: {len(reader.pages)}", ""]
    for page_num, page in enumerate(reader.pages, 1):
        text = page.extract_text() or ""
        text = re.sub(r"[ \t]+", " ", text).strip()
        if text:
            lines.extend([f"## Page {page_num}", "", text, ""])
    return "\n".join(lines).rstrip() + "\n"


def extract_docx(path: Path) -> str:
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    lines = [f"# Extracted Text: {path.name}", ""]
    with zipfile.ZipFile(path) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    for para in root.findall(".//w:p", ns):
        bits = []
        for node in para.findall(".//w:t", ns):
            if node.text:
                bits.append(node.text)
        text = clean(unescape("".join(bits)))
        if text:
            lines.append(text)
    return "\n\n".join(lines).rstrip() + "\n"


def extract_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=False)
    lines = [f"# Extracted Workbook Preview: {path.name}", ""]
    for sheet in workbook.worksheets:
        lines.extend([f"## Sheet: {sheet.title}", ""])
        for row in sheet.iter_rows(values_only=True):
            values = [clean(cell) for cell in row]
            while values and not values[-1]:
                values.pop()
            if not any(values):
                continue
            lines.append("| " + " | ".join(value.replace("|", "\\|") or "-" for value in values) + " |")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def ensure_extract(source: dict[str, str], source_path: Path, *, write: bool) -> str | None:
    extract_rel = source.get("extract")
    if not extract_rel:
        return None
    extract_path = SOURCE_DIR / extract_rel
    if not write:
        if not extract_path.is_file():
            raise SystemExit(f"Missing source extract: {extract_path.relative_to(ROOT)}")
        return extract_rel
    kind = source["kind"]
    if kind == "pdf":
        content = extract_pdf(source_path)
    elif kind == "docx":
        content = extract_docx(source_path)
    elif kind == "xlsx":
        content = extract_xlsx(source_path)
    else:
        raise SystemExit(f"Unsupported extract kind {kind!r} for {source_path}")
    write_if_changed(extract_path, content)
    return extract_rel


def recipe_ref_from_item(item: dict[str, Any]) -> str | None:
    for key in ("recipeId", "recipe_id", "id", "builtin_id"):
        value = item.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return None


def build_package_map(packages: dict[str, Any]) -> dict[str, list[str]]:
    mapping: dict[str, list[str]] = defaultdict(list)
    for package in packages.get("packages", []):
        package_id = clean(package.get("id"))
        package_label = "Riviera Table / Offsite" if package_id == "offsite" else clean(package.get("label") or package_id)
        for section in package.get("sections", []) or []:
            section_label = clean(section.get("label") or section.get("id"))
            for course in section.get("courses", []) or []:
                for item in course.get("items", []) or []:
                    if not isinstance(item, dict):
                        continue
                    recipe_id = recipe_ref_from_item(item)
                    if not recipe_id:
                        continue
                    ref = f"{package_label}: {section_label}" if section_label else package_label
                    if ref not in mapping[recipe_id]:
                        mapping[recipe_id].append(ref)
    return dict(mapping)


def md_escape(text: Any) -> str:
    value = clean(text)
    return value.replace("|", "\\|")


def recipe_tags(recipe: dict[str, Any], package_refs: list[str]) -> list[str]:
    tags = ["House Standard"]
    for key in ("course", "type", "method"):
        value = clean(recipe.get(key))
        if value and value not in tags:
            tags.append(value)
    for key in ("diet", "protein"):
        for value in recipe.get(key) or []:
            value = clean(value)
            if value and value not in tags:
                tags.append(value)
    for ref in package_refs:
        package = ref.split(":", 1)[0]
        if package and package not in tags:
            tags.append(package)
    return tags


def recipe_to_markdown(recipe: dict[str, Any], package_refs: list[str]) -> str:
    lines: list[str] = []
    lines.append(f"### {md_escape(recipe.get('name'))}")
    lines.append("")
    lines.append(f"- Recipe ID: `{recipe['id']}`")
    lines.append(f"- Status: {md_escape(recipe.get('status') or '-')}")
    lines.append(f"- Version: {md_escape(recipe.get('version') or '-')}")
    lines.append(f"- Yield: {md_escape(recipe.get('yield') or '-')}")
    lines.append(f"- Label: {md_escape(recipe.get('label') or '-')}")
    provenance = recipe.get("provenance") if isinstance(recipe.get("provenance"), dict) else {}
    if provenance:
        source = clean(provenance.get("source"))
        source_date = clean(provenance.get("sourceDate"))
        scope = clean(provenance.get("scope"))
        detail = " | ".join(value for value in (source, source_date, scope) if value)
        lines.append(f"- Provenance: {md_escape(detail)}")
    lines.append(f"- Tags: {md_escape(' | '.join(recipe_tags(recipe, package_refs)))}")
    aliases = [clean(alias) for alias in recipe.get("aliases") or [] if clean(alias)]
    if aliases:
        lines.append(f"- Aliases: {md_escape(' | '.join(aliases))}")
    if package_refs:
        lines.append(f"- Package / section refs: {md_escape('; '.join(package_refs))}")
    subtitle = clean(recipe.get("subtitle"))
    if subtitle:
        lines.append(f"- Menu description: {md_escape(subtitle)}")
    elements = [clean(x) for x in recipe.get("elements") or [] if clean(x)]
    if elements:
        lines.append(f"- Elements: {md_escape(' | '.join(elements))}")
    lines.append("")

    confirmation_flags = [
        clean(flag) for flag in recipe.get("confirmationFlags") or [] if clean(flag)
    ]
    if confirmation_flags:
        lines.append("#### Needs Confirmation")
        lines.append("")

    allergens = recipe.get("allergens") if isinstance(recipe.get("allergens"), dict) else {}
    if allergens:
        lines.append("#### Allergens")
        lines.append("")
        lines.append(f"- Status: {md_escape(allergens.get('status') or '-')}")
        lines.append(
            f"- Contains: {md_escape(' | '.join(allergens.get('contains') or []) or 'None confirmed')}"
        )
        lines.append(
            f"- May contain: {md_escape(' | '.join(allergens.get('mayContain') or []) or 'None confirmed')}"
        )
        if clean(allergens.get("notes")):
            lines.append(f"- Notes: {md_escape(allergens.get('notes'))}")
        lines.append("")

    controls = recipe.get("controls") if isinstance(recipe.get("controls"), dict) else {}
    if controls:
        lines.append("#### Operational Controls")
        lines.append("")

    scaling = recipe.get("scalingBasis") if isinstance(recipe.get("scalingBasis"), dict) else {}
    if scaling:
        lines.append("#### Scaling Basis")
        lines.append("")
        lines.append(f"- Status: {md_escape(scaling.get('status') or '-')}")
        lines.append(f"- Basis: {md_escape(scaling.get('basis') or 'Not confirmed')}")
        lines.append(f"- Base yield: {md_escape(scaling.get('baseYield') or '-')}")
        if clean(scaling.get("notes")):
            lines.append(f"- Notes: {md_escape(scaling.get('notes'))}")
        lines.append("")

    rational = (
        recipe.get("rationalSettings")
        if isinstance(recipe.get("rationalSettings"), dict)
        else {}
    )
    if rational:
        lines.append("#### Rational Settings")
        lines.append("")
        lines.append(f"- Status: {md_escape(rational.get('status') or '-')}")
        for stage in rational.get("stages") or []:
            detail = " | ".join(
                f"{key}={value}"
                for key, value in stage.items()
                if value not in (None, "")
            )
            lines.append(f"- {md_escape(detail)}")
        if not rational.get("stages"):
            lines.append("- No confirmed stages.")
        if clean(rational.get("notes")):
            lines.append(f"- Notes: {md_escape(rational.get('notes'))}")
        lines.append("")
        for control_key in ("cooling", "holding", "packing", "service"):
            control = controls.get(control_key) if isinstance(controls.get(control_key), dict) else {}
            lines.append(
                f"- {control_key.title()}: {md_escape(control.get('status') or 'NEEDS CONFIRMATION')}"
            )
            for step in control.get("steps") or []:
                lines.append(f"  - {md_escape(step)}")
        lines.append("")
        for flag in confirmation_flags:
            lines.append(f"- {md_escape(flag)}")
        lines.append("")

    ingredients = recipe.get("ingredients") or []
    if ingredients:
        lines.append("#### Ingredients")
        lines.append("")
        lines.append("| Qty | Ingredient / prep |")
        lines.append("| --- | --- |")
        for ing in ingredients:
            item = clean(ing.get("item"))
            prep = clean(ing.get("prep"))
            detail = item if not prep else f"{item} - {prep}"
            lines.append(f"| {md_escape(ing.get('qty') or '-')} | {md_escape(detail)} |")
        lines.append("")

    method_steps = [clean(x) for x in recipe.get("method_steps") or [] if clean(x)]
    if method_steps:
        lines.append("#### Prep method")
        lines.append("")
        for idx, step in enumerate(method_steps, 1):
            lines.append(f"{idx}. {step}")
        lines.append("")

    service_steps = [clean(x) for x in recipe.get("service") or [] if clean(x)]
    if service_steps:
        lines.append("#### On the day / service")
        lines.append("")
        for idx, step in enumerate(service_steps, 1):
            lines.append(f"{idx}. {step}")
        lines.append("")

    note = clean(recipe.get("note"))
    if note:
        lines.append("#### Notes")
        lines.append("")
        lines.append(note)
        lines.append("")
    return "\n".join(lines).strip() + "\n"


def build_tapas_overlay(recipes: list[dict[str, Any]], packages: dict[str, Any]) -> str:
    by_id = {str(recipe["id"]): recipe for recipe in recipes}
    missing = [recipe_id for recipe_id in JULY_8_OVERLAY_IDS if recipe_id not in by_id]
    if missing:
        raise SystemExit(f"Missing house-standard recipes: {missing}")

    package_map = build_package_map(packages)
    lines: list[str] = [
        "# Riviera Tapas House Standards Overlay 2026-07-08",
        "",
        "**Status:** Active GitHub recipe-data overlay with historical ChatGPT source provenance.",
        "**Source:** `Tapas Canape Recipe Cards.docx`, standardised into the structured Riviera recipe catalog on 2026-07-08.",
        "**Use for:** House-standard tapas/canape recipe cards, kitchen PDFs, prep sheets, Sunday tapas planning, and package-linked canape/tapas pulls.",
        "",
        "## Authority",
        "",
        "These 16 recipes override older ChatGPT project recipe-bank versions for the same dishes. The structured GitHub catalog controls current recipe data. Approved Google Drive masters control operational SOPs, packages, live orders and supplier data; ChatGPT receives those domains through a read-optimised published release.",
        "",
        "## Overlay Recipe Index",
        "",
        "| Recipe ID | Recipe | Yield | Package refs |",
        "| --- | --- | --- | --- |",
    ]
    for recipe_id in JULY_8_OVERLAY_IDS:
        recipe = by_id[recipe_id]
        refs = package_map.get(recipe_id, [])
        lines.append(
            f"| `{recipe_id}` | {md_escape(recipe.get('name'))} | {md_escape(recipe.get('yield') or '-')} | {md_escape('; '.join(refs) or 'Standalone / base')} |"
        )
    lines.extend(["", "## Full Overlay Recipes", ""])
    for recipe_id in JULY_8_OVERLAY_IDS:
        lines.append(recipe_to_markdown(by_id[recipe_id], package_map.get(recipe_id, [])))
    return "\n".join(lines).rstrip() + "\n"


def build_current_house_standard_additions(recipes: list[dict[str, Any]], packages: dict[str, Any]) -> str:
    by_id = {str(recipe["id"]): recipe for recipe in recipes}
    missing = [recipe_id for recipe_id in CURRENT_HOUSE_STANDARD_IDS if recipe_id not in by_id]
    if missing:
        raise SystemExit(f"Missing current house-standard recipes: {missing}")

    marked = [
        str(recipe.get("id"))
        for recipe in recipes
        if isinstance(recipe, dict) and recipe.get("houseStandard") is True
    ]
    if marked != CURRENT_HOUSE_STANDARD_IDS:
        raise SystemExit(f"Current house-standard order drifted: {marked}")

    addition_ids = [recipe_id for recipe_id in CURRENT_HOUSE_STANDARD_IDS if recipe_id not in JULY_8_OVERLAY_IDS]
    package_map = build_package_map(packages)
    lines: list[str] = [
        "# Riviera Current House Standards Additions 2026-07-27",
        "",
        "**Status:** Active direct user-approved house standards added after the fixed July 8 overlay.",
        "**Authority:** These additions are canonical in the structured GitHub recipe catalog and outrank historical ChatGPT recipe data for the listed recipes.",
        "",
        "## Addition Index",
        "",
        "| Recipe ID | Recipe | Yield | Package refs |",
        "| --- | --- | --- | --- |",
    ]
    for recipe_id in addition_ids:
        recipe = by_id[recipe_id]
        refs = package_map.get(recipe_id, [])
        lines.append(
            f"| `{recipe_id}` | {md_escape(recipe.get('name'))} | {md_escape(recipe.get('yield') or '-')} | {md_escape('; '.join(refs) or 'Standalone / base')} |"
        )
    lines.extend(["", "## Full Recipes", ""])
    for recipe_id in addition_ids:
        lines.append(recipe_to_markdown(by_id[recipe_id], package_map.get(recipe_id, [])))
    return "\n".join(lines).rstrip() + "\n"


def source_inventory(*, write_extracts: bool) -> list[dict[str, Any]]:
    rows = []
    for source in SOURCE_FILES:
        filename = source["file"]
        path = SOURCE_DIR / filename
        if not path.exists():
            raise SystemExit(f"Missing ChatGPT source file: {path}")
        extract_rel = ensure_extract(source, path, write=write_extracts)
        text_path = SOURCE_DIR / (extract_rel or filename)
        text: str | None = None
        lines: int | None = None
        if text_path.suffix.lower() in {".md", ".txt"}:
            text = text_path.read_text(encoding="utf-8")
            lines = line_count(text)
        rows.append(
            {
                "file": filename,
                "liveName": source.get("liveName", filename),
                "kind": source["kind"],
                "bytes": path.stat().st_size,
                "lines": lines,
                "sha256": sha256(path),
                "extract": extract_rel,
                "extractLines": line_count(text) if text is not None and extract_rel else None,
            }
        )
    return rows


def build_source_of_truth(
    overlay: str,
    current_house_standard_additions: str,
    inventory: list[dict[str, Any]],
) -> str:
    lines: list[str] = [
        "# Riviera Source Of Truth 2026-07-08",
        "",
        "**Status:** Active GitHub recipe-data bundle with historical source appendices.",
        f"**Active release ID:** `{ACTIVE_RELEASE_ID}`.",
        "**Merge direction:** The structured GitHub catalog is canonical for recipe data. The fixed July 8 tapas/canape overlay and later approved standards supersede historical ChatGPT recipe versions; approved Drive masters remain canonical for operations and ChatGPT receives a read-optimised release.",
        "**Generated by:** `scripts/build_riviera_source_of_truth.py`.",
        "",
        "## Authority And Conflict Rules",
        "",
        "1. A direct user correction controls the current task, but is not a published source update until it is scope-classified, recorded in a change receipt, applied to the correct master, republished and verified.",
        "2. Approved Google Drive masters control operational SOPs, package rules, live orders, approvals, prices, stock, delivery requirements and archives.",
        "3. `riviera_sources/current/Riviera_Recipe_Catalog_Source_Of_Truth_2026-07-08.json` is the canonical structured recipe payload.",
        "4. The July 8 Tapas House Standards Overlay in this file outranks older ChatGPT recipe-bank content for the same 16 dishes.",
        "5. Later approved house standards recorded in the structured catalog outrank historical ChatGPT recipe data for those recipes; they do not rewrite the fixed historical overlay.",
        "6. The ChatGPT source pack downloaded on 2026-07-08 is historical provenance for inherited rules, not the current operational master. ChatGPT daily work uses a read-optimised published release.",
        "7. `riviera_data/builtins.json` and generated PDFs are recipe-data representations. Package JSON must be reconciled to the approved Drive package master before cross-system publication.",
        "8. Do not silently invent missing recipe, package, dietary, ordering, or service rules. Mark `NEEDS CONFIRMATION` when sources conflict or a required detail is absent.",
        "9. Temporary event/week instructions require an expiry and must not be promoted into a permanent SOP.",
        "",
        "## Active Source Stack",
        "",
        "| Order | Live project source | Stored file | Kind | Lines | SHA-256 |",
        "| ---: | --- | --- | --- | ---: | --- |",
    ]
    for idx, row in enumerate(inventory, 1):
        row_lines = row["lines"] if row["lines"] is not None else "-"
        lines.append(
            f"| {idx} | {md_escape(row['liveName'])} | `{row['file']}` | {md_escape(row['kind'])} | {row_lines} | `{row['sha256'][:16]}` |"
        )
    lines.extend(
        [
            "",
            "## July 8 Tapas Overlay",
            "",
            "The following section is embedded from `Riviera_Tapas_House_Standards_Overlay_2026-07-08.md`.",
            "",
            overlay.strip(),
            "",
            "## Current House Standards Added After July 8",
            "",
            "The following section is embedded from `Riviera_Current_House_Standards_Additions_2026-07-27.md`.",
            "",
            current_house_standard_additions.strip(),
            "",
            "## Historical ChatGPT Source Appendices",
            "",
            "The appendices below preserve the 2026-07-08 ChatGPT Riviera Project snapshot as historical provenance.",
            "",
        ]
    )
    for row in inventory:
        content_rel = row.get("extract") or row["file"]
        content_path = SOURCE_DIR / content_rel
        content = content_path.read_text(encoding="utf-8").strip()
        lines.extend(
            [
                f"### Source: {row['liveName']}",
                "",
                f"- Stored file: `{row['file']}`",
                f"- Kind: {row['kind']}",
                f"- Lines: {row['lines']}",
                f"- SHA-256: `{row['sha256']}`",
            "",
            ]
        )
        if row.get("extract"):
            lines.append(f"- Text extract: `{row['extract']}`")
            lines.append("")
        else:
            lines.append("")
        lines.extend(
            [
                "````markdown",
                content,
                "````",
                "",
            ]
        )
    return "\n".join(lines).rstrip() + "\n"


def expected_manifest(inventory: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "status": "active",
        "date": "2026-07-27",
        "releaseId": ACTIVE_RELEASE_ID,
        "mergeDirection": "GitHub is canonical for structured recipe data; the July 8 overlay and later approved standards supersede historical ChatGPT recipe versions. Drive remains canonical for operations and ChatGPT receives a read-optimised release.",
        "sourceOfTruth": str(SOURCE_OF_TRUTH_PATH.relative_to(ROOT)),
        "structuredRecipeCatalog": str(RECIPE_CATALOG_PATH.relative_to(ROOT)),
        "overlay": str(TAPAS_OVERLAY_PATH.relative_to(ROOT)),
        "currentHouseStandardsAdditions": str(CURRENT_STANDARDS_ADDITIONS_PATH.relative_to(ROOT)),
        "chatgptSourceDir": str(SOURCE_DIR.relative_to(ROOT)),
        "houseStandardOverlayRecipeIds": JULY_8_OVERLAY_IDS,
        "currentHouseStandardRecipeIds": CURRENT_HOUSE_STANDARD_IDS,
        "chatgptSources": inventory,
    }


def check_generated_file(path: Path, expected: str, errors: list[str]) -> None:
    if not path.is_file():
        errors.append(f"missing generated file: {path.relative_to(ROOT)}")
        return
    actual = path.read_text(encoding="utf-8")
    if actual != expected:
        errors.append(
            f"generated drift: {path.relative_to(ROOT)}; "
            "run python3 scripts/build_riviera_source_of_truth.py --write"
        )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--check", action="store_true", help="verify generated SSOT files without writing")
    mode.add_argument("--write", action="store_true", help="rebuild generated SSOT files")
    args = parser.parse_args()

    write = not args.check
    recipes = load_recipe_catalog()
    packages = load_json(PACKAGES_PATH)
    inventory = source_inventory(write_extracts=write)
    overlay = build_tapas_overlay(recipes, packages)
    current_house_standard_additions = build_current_house_standard_additions(recipes, packages)
    source_of_truth = build_source_of_truth(overlay, current_house_standard_additions, inventory)
    manifest_text = json.dumps(expected_manifest(inventory), indent=2) + "\n"

    if args.check:
        errors: list[str] = []
        check_generated_file(TAPAS_OVERLAY_PATH, overlay, errors)
        check_generated_file(CURRENT_STANDARDS_ADDITIONS_PATH, current_house_standard_additions, errors)
        check_generated_file(SOURCE_OF_TRUTH_PATH, source_of_truth, errors)
        check_generated_file(MANIFEST_PATH, manifest_text, errors)
        if errors:
            for error in errors:
                print(f"ERROR: {error}", file=sys.stderr)
            return 1
        print(
            json.dumps(
                {
                    "status": "ok",
                    "mode": "check",
                    "sourceOfTruth": str(SOURCE_OF_TRUTH_PATH.relative_to(ROOT)),
                    "sources": len(inventory),
                    "overlayRecipes": len(JULY_8_OVERLAY_IDS),
                    "currentHouseStandards": len(CURRENT_HOUSE_STANDARD_IDS),
                },
                indent=2,
            )
        )
        return 0

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    write_if_changed(TAPAS_OVERLAY_PATH, overlay)
    write_if_changed(CURRENT_STANDARDS_ADDITIONS_PATH, current_house_standard_additions)
    write_if_changed(SOURCE_OF_TRUTH_PATH, source_of_truth)
    write_if_changed(MANIFEST_PATH, manifest_text)
    print(
        json.dumps(
            {
                "status": "ok",
                "mode": "write",
                "sourceOfTruth": str(SOURCE_OF_TRUTH_PATH.relative_to(ROOT)),
                "sources": len(inventory),
                "overlayRecipes": len(JULY_8_OVERLAY_IDS),
                "currentHouseStandards": len(CURRENT_HOUSE_STANDARD_IDS),
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
